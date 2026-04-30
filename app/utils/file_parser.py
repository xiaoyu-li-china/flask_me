import json
import os
import zipfile
from datetime import datetime
from xml.etree import ElementTree as ET

try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

try:
    import pandas as pd
    HAS_PANDAS = True
except ImportError:
    HAS_PANDAS = False


class BaseFileParser:
    def __init__(self, file_path):
        self.file_path = file_path
        self.file_name = os.path.basename(file_path)
    
    def parse(self):
        raise NotImplementedError("Subclasses must implement parse method")
    
    def get_file_extension(self):
        return os.path.splitext(self.file_name)[1].lower()


class ExcelParser(BaseFileParser):
    def parse(self):
        if not HAS_OPENPYXL:
            raise ImportError("openpyxl is required for Excel parsing. Please install it with 'pip install openpyxl'")
        
        try:
            workbook = openpyxl.load_workbook(self.file_path, read_only=True, data_only=True)
            parsed_data = {
                'file_name': self.file_name,
                'file_type': 'excel',
                'sheets': []
            }
            
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                sheet_data = {
                    'sheet_name': sheet_name,
                    'headers': [],
                    'rows': []
                }
                
                rows = list(sheet.iter_rows(values_only=True))
                if len(rows) > 0:
                    sheet_data['headers'] = [str(cell) if cell is not None else '' for cell in rows[0]]
                    
                    for row in rows[1:]:
                        if any(cell is not None and cell != '' for cell in row):
                            row_dict = {}
                            for i, cell in enumerate(row):
                                header = sheet_data['headers'][i] if i < len(sheet_data['headers']) else f'Column_{i}'
                                row_dict[header] = str(cell) if cell is not None else ''
                            sheet_data['rows'].append(row_dict)
                
                parsed_data['sheets'].append(sheet_data)
            
            workbook.close()
            return self._format_test_cases(parsed_data)
        
        except Exception as e:
            raise Exception(f"Error parsing Excel file: {str(e)}")
    
    def _format_test_cases(self, parsed_data):
        test_cases = []
        
        for sheet in parsed_data['sheets']:
            for row in sheet['rows']:
                test_case = {}
                
                title_keys = ['用例标题', '测试用例', '标题', 'Title', 'Test Case', 'Case Name']
                for key in title_keys:
                    if key in row and row[key]:
                        test_case['title'] = row[key]
                        break
                
                if 'title' not in test_case:
                    for key, value in row.items():
                        if value and key not in ['序号', '编号', 'ID']:
                            test_case['title'] = value
                            break
                
                if 'title' not in test_case:
                    continue
                
                description_keys = ['描述', '说明', '备注', 'Description', 'Notes', 'Remark']
                for key in description_keys:
                    if key in row and row[key]:
                        test_case['description'] = row[key]
                        break
                
                steps_keys = ['测试步骤', '步骤', '操作步骤', 'Steps', 'Test Steps', 'Action']
                for key in steps_keys:
                    if key in row and row[key]:
                        test_case['test_steps'] = row[key]
                        break
                
                expected_keys = ['预期结果', '期望结果', 'Expected', 'Expected Result']
                for key in expected_keys:
                    if key in row and row[key]:
                        test_case['expected_result'] = row[key]
                        break
                
                priority_keys = ['优先级', 'Priority', 'P0', 'P1', 'P2']
                for key in priority_keys:
                    if key in row and row[key]:
                        test_case['priority'] = row[key]
                        break
                
                module_keys = ['模块', 'Module', '所属模块']
                for key in module_keys:
                    if key in row and row[key]:
                        test_case['module'] = row[key]
                        break
                
                sub_module_keys = ['子模块', 'SubModule', 'Sub Module']
                for key in sub_module_keys:
                    if key in row and row[key]:
                        test_case['sub_module'] = row[key]
                        break

                test_case_id_keys = ['TestCaseId', 'Test Case ID', '用例ID', '用例编号', 'Case ID']
                for key in test_case_id_keys:
                    if key in row and row[key]:
                        test_case['test_case_id'] = row[key]
                        break

                preconditions_keys = ['Preconditions', '前置条件', '前提条件']
                for key in preconditions_keys:
                    if key in row and row[key]:
                        test_case['preconditions'] = row[key]
                        break

                category_keys = ['分类', 'Category', '测试类型']
                for key in category_keys:
                    if key in row and row[key]:
                        test_case['category'] = row[key]
                        break
                
                test_case['raw_data'] = row
                test_case['source_sheet'] = sheet['sheet_name']
                
                test_cases.append(test_case)
        
        result = {
            'file_name': self.file_name,
            'file_type': 'excel',
            'total_cases': len(test_cases),
            'test_cases': test_cases,
            'sheets': parsed_data.get('sheets', [])
        }
        
        return result


class XMindParser(BaseFileParser):
    def parse(self):
        try:
            if not zipfile.is_zipfile(self.file_path):
                raise Exception("Invalid XMind file. XMind files should be ZIP archives.")
            
            with zipfile.ZipFile(self.file_path, 'r') as zf:
                content_xml = None
                manifest_xml = None
                
                for name in zf.namelist():
                    if name.endswith('content.xml'):
                        content_xml = name
                    if name.endswith('manifest.json'):
                        manifest_json = name
                
                if content_xml:
                    return self._parse_xmind_content_xml(zf, content_xml)
                elif manifest_json:
                    return self._parse_xmind_manifest_json(zf, manifest_json)
                else:
                    raise Exception("Cannot find content.xml or manifest.json in XMind file")
        
        except Exception as e:
            raise Exception(f"Error parsing XMind file: {str(e)}")
    
    def _parse_xmind_content_xml(self, zf, content_xml_path):
        with zf.open(content_xml_path) as f:
            content = f.read()
        
        root = ET.fromstring(content)
        
        test_cases = []
        
        ns = {'xmap': 'urn:xmind:xmap:xmlns:content:2.0'}
        
        sheets = root.findall('.//xmap:sheet', ns)
        if not sheets:
            sheets = root.findall('.//sheet')
        
        for sheet in sheets:
            sheet_title = sheet.get('title', 'Untitled Sheet')
            
            topics = sheet.findall('.//xmap:topic', ns)
            if not topics:
                topics = sheet.findall('.//topic')
            
            for topic in topics:
                test_case = self._parse_topic_to_test_case(topic, sheet_title)
                if test_case:
                    test_cases.append(test_case)
        
        result = {
            'file_name': self.file_name,
            'file_type': 'xmind',
            'total_cases': len(test_cases),
            'test_cases': test_cases
        }
        
        return result
    
    def _parse_topic_to_test_case(self, topic, sheet_title):
        ns = {'xmap': 'urn:xmind:xmap:xmlns:content:2.0'}
        
        title_elem = topic.find('xmap:title', ns)
        if title_elem is None:
            title_elem = topic.find('title')
        
        title = title_elem.text if title_elem is not None and title_elem.text else 'Untitled'
        
        test_case = {
            'title': title,
            'module': sheet_title,
            'source_sheet': sheet_title
        }
        
        children = topic.findall('.//xmap:children/xmap:topics/xmap:topic', ns)
        if not children:
            children = topic.findall('.//children/topics/topic')
        
        steps = []
        expected_results = []
        
        for child in children:
            child_title_elem = child.find('xmap:title', ns)
            if child_title_elem is None:
                child_title_elem = child.find('title')
            
            child_title = child_title_elem.text if child_title_elem is not None and child_title_elem.text else ''
            
            child_text_lower = child_title.lower()
            
            if '步骤' in child_text_lower or 'step' in child_text_lower:
                steps.append(child_title)
            elif '预期' in child_text_lower or 'expected' in child_text_lower:
                expected_results.append(child_title)
            elif '描述' in child_text_lower or 'description' in child_text_lower:
                test_case['description'] = child_title
            elif '优先级' in child_text_lower or 'priority' in child_text_lower:
                test_case['priority'] = child_title
            else:
                steps.append(child_title)
        
        if steps:
            test_case['test_steps'] = '\n'.join([f"{i+1}. {step}" for i, step in enumerate(steps)])
        
        if expected_results:
            test_case['expected_result'] = '\n'.join(expected_results)
        
        return test_case
    
    def _parse_xmind_manifest_json(self, zf, manifest_json_path):
        import json
        
        with zf.open(manifest_json_path) as f:
            manifest = json.load(f)
        
        test_cases = []
        
        sheets = manifest.get('sheets', [])
        for sheet in sheets:
            sheet_title = sheet.get('title', 'Untitled Sheet')
            
            root_topic = sheet.get('rootTopic', {})
            if root_topic:
                test_cases_from_topic = self._parse_json_topic(root_topic, sheet_title)
                test_cases.extend(test_cases_from_topic)
        
        result = {
            'file_name': self.file_name,
            'file_type': 'xmind',
            'total_cases': len(test_cases),
            'test_cases': test_cases
        }
        
        return result
    
    def _parse_json_topic(self, topic, sheet_title, parent_path=None):
        test_cases = []
        
        if parent_path is None:
            parent_path = []
        
        title = topic.get('title', 'Untitled')
        current_path = parent_path + [title]
        
        children = topic.get('children', {}).get('topics', [])
        
        if not children:
            if len(current_path) >= 2:
                test_case = {
                    'title': ' - '.join(current_path[-2:]),
                    'module': sheet_title,
                    'source_sheet': sheet_title,
                    'description': ' -> '.join(current_path)
                }
                test_cases.append(test_case)
        else:
            for child in children:
                child_cases = self._parse_json_topic(child, sheet_title, current_path)
                test_cases.extend(child_cases)
        
        return test_cases


def parse_file(file_path, file_extension=None):
    file_name = os.path.basename(file_path)
    
    if file_extension:
        ext = file_extension.lower()
    else:
        ext = os.path.splitext(file_name)[1].lower()
    
    if not ext.startswith('.'):
        ext = '.' + ext
    
    if ext in ['.xlsx', '.xls']:
        parser = ExcelParser(file_path)
    elif ext in ['.xmind']:
        parser = XMindParser(file_path)
    else:
        raise Exception(f"Unsupported file format: {ext}")
    
    return parser.parse()
